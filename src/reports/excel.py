import logging
from datetime import datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.reconciliation.models import ReconciliationRow

logger = logging.getLogger(__name__)

# ── Styles ────────────────────────────────────────────────────────────────────
_RED_FILL    = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
_GREEN_FILL  = PatternFill(start_color="FF92D050", end_color="FF92D050", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")
_SECT_FILL   = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFFFF", bold=True, size=11)
_BOLD_FONT   = Font(bold=True)
_TITLE_FONT  = Font(bold=True, size=13)
_AMOUNT_FMT  = "#,##0"       # CLP — sin decimales
_HH_FMT      = "#,##0.00"    # horas — dos decimales
_CENTER      = Alignment(horizontal="center", vertical="center")
_LEFT        = Alignment(horizontal="left",   vertical="center")
_RIGHT       = Alignment(horizontal="right",  vertical="center")

# ── Column definitions ────────────────────────────────────────────────────────
# Metadata columns (common to both sheet types)
_META_COLS: list[tuple[str, int]] = [
    ("Periodo",       12),
    ("Sitio",         30),
    ("N° Reporte",    13),
    ("Estado",        18),
    ("RUT",           12),
    ("Nombre",        32),
    ("Cargo",         25),
    ("Cód. Haber",    13),
    ("Descripción",   22),
]
_N_META = len(_META_COLS)

# Numeric columns differ between haberes (CLP) and horas extras (HH)
_AMOUNT_COLS: list[tuple[str, int]] = [
    ("Monto Oracle",  15),
    ("Monto Buk",     13),
    ("Diferencia",    13),
]
_HH_COLS: list[tuple[str, int]] = [
    ("HH Oracle",     13),
    ("HH Buk",        11),
    ("Diferencia HH", 15),
]

_ALL_HABERES_COLS = _META_COLS + _AMOUNT_COLS
_ALL_HH_COLS      = _META_COLS + _HH_COLS
_N_TOTAL          = len(_ALL_HABERES_COLS)  # same for both (12)


def _is_haber(rec: ReconciliationRow) -> bool:
    return rec.cod_haber.startswith("B")


def _is_he(rec: ReconciliationRow) -> bool:
    return rec.cod_haber.startswith("H")


# ── Generic helpers ────────────────────────────────────────────────────────────

def _set_header_row(ws, columns: list[tuple[str, int]], row: int = 1) -> None:
    for col_idx, (label, width) in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=label)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTER
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_data_row(
    ws,
    row_num: int,
    rec: ReconciliationRow,
    num_fmt: str,
) -> None:
    meta_values = [
        rec.periodo,
        rec.nom_sitio,
        rec.num_reporte,
        rec.estado_reporte,
        rec.rut,
        rec.nombre,
        rec.cargo,
        rec.cod_haber,
        rec.descripcion_haber,
    ]
    num_values = [
        float(rec.monto_oracle),
        float(rec.monto_buk),
        float(rec.diferencia),
    ]
    fill = _RED_FILL if rec.tiene_diferencia else None

    for col_idx, value in enumerate(meta_values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.alignment = _LEFT
        if fill:
            cell.fill = fill

    for offset, value in enumerate(num_values):
        col_idx = _N_META + 1 + offset
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.number_format = num_fmt
        cell.alignment     = _RIGHT
        if fill:
            cell.fill = fill


def _write_totals_row(ws, row_num: int, rows: list[ReconciliationRow], num_fmt: str) -> None:
    total_oracle = float(sum(r.monto_oracle for r in rows))
    total_buk    = float(sum(r.monto_buk    for r in rows))
    total_diff   = float(sum(r.diferencia   for r in rows))

    ws.cell(row=row_num, column=_N_META, value="TOTAL").font = _BOLD_FONT
    for offset, val in enumerate([total_oracle, total_buk, total_diff]):
        col_idx = _N_META + 1 + offset
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.number_format = num_fmt
        cell.font          = _BOLD_FONT
        cell.alignment     = _RIGHT


def _freeze_and_filter(ws, n_cols: int) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}1"


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _build_haberes_sheet(wb: Workbook, rows: list[ReconciliationRow]) -> None:
    """Sheet 'Haberes' — códigos B%, montos en CLP."""
    ws   = wb.create_sheet("Haberes")
    data = [r for r in rows if _is_haber(r)]
    _set_header_row(ws, _ALL_HABERES_COLS)
    for i, rec in enumerate(data, start=2):
        _write_data_row(ws, i, rec, _AMOUNT_FMT)
    if data:
        _write_totals_row(ws, len(data) + 2, data, _AMOUNT_FMT)
    _freeze_and_filter(ws, _N_TOTAL)


def _build_he_sheet(wb: Workbook, rows: list[ReconciliationRow]) -> None:
    """Sheet 'Horas Extras HH' — códigos H%, valores en horas."""
    ws   = wb.create_sheet("Horas Extras HH")
    data = [r for r in rows if _is_he(r)]
    _set_header_row(ws, _ALL_HH_COLS)
    for i, rec in enumerate(data, start=2):
        _write_data_row(ws, i, rec, _HH_FMT)
    if data:
        _write_totals_row(ws, len(data) + 2, data, _HH_FMT)
    _freeze_and_filter(ws, _N_TOTAL)


def _build_diff_sheet(wb: Workbook, rows: list[ReconciliationRow]) -> None:
    """Sheet 'Solo Diferencias' — filas con diferencia de ambos tipos."""
    diff_rows = [r for r in rows if r.tiene_diferencia]
    ws = wb.create_sheet("Solo Diferencias")
    if not diff_rows:
        cell = ws.cell(row=1, column=1, value="Sin diferencias encontradas")
        cell.fill = _GREEN_FILL
        cell.font = _BOLD_FONT
        return

    haberes_diff = [r for r in diff_rows if _is_haber(r)]
    he_diff      = [r for r in diff_rows if _is_he(r)]
    row_num = 1

    # ── Sección Haberes (B%) ──────────────────────────────────────────────────
    if haberes_diff:
        title = ws.cell(row=row_num, column=1, value="Haberes con Diferencia")
        title.font = _BOLD_FONT
        title.fill = _SECT_FILL
        title.font = Font(color="FFFFFFFF", bold=True)
        row_num += 1
        _set_header_row(ws, _ALL_HABERES_COLS, row=row_num)
        row_num += 1
        for rec in haberes_diff:
            _write_data_row(ws, row_num, rec, _AMOUNT_FMT)
            row_num += 1
        _write_totals_row(ws, row_num, haberes_diff, _AMOUNT_FMT)
        row_num += 2

    # ── Sección Horas Extras (H%) ─────────────────────────────────────────────
    if he_diff:
        title = ws.cell(row=row_num, column=1, value="Horas Extras HH con Diferencia")
        title.font = Font(color="FFFFFFFF", bold=True)
        title.fill = _SECT_FILL
        row_num += 1
        _set_header_row(ws, _ALL_HH_COLS, row=row_num)
        row_num += 1
        for rec in he_diff:
            _write_data_row(ws, row_num, rec, _HH_FMT)
            row_num += 1
        _write_totals_row(ws, row_num, he_diff, _HH_FMT)

    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30


def _build_summary_sheet(wb: Workbook, rows: list[ReconciliationRow], periodo: str) -> None:
    ws = wb.create_sheet("Resumen")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    for col in ["C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 16

    title = ws.cell(row=1, column=1, value=f"Resumen Cuadratura — Período {periodo}")
    title.font = _TITLE_FONT

    haberes_rows = [r for r in rows if _is_haber(r)]
    he_rows      = [r for r in rows if _is_he(r)]

    row = 3
    for section_label, section_rows, num_fmt in [
        ("Haberes (B%) — Montos CLP", haberes_rows, _AMOUNT_FMT),
        ("Horas Extras (H%) — HH",   he_rows,      _HH_FMT),
    ]:
        if not section_rows:
            continue

        # Section title
        sec = ws.cell(row=row, column=1, value=section_label)
        sec.font = Font(color="FFFFFFFF", bold=True)
        sec.fill = _SECT_FILL
        row += 1

        # By sitio
        ws.cell(row=row, column=1, value="Por Sitio").font = _BOLD_FONT
        row += 1
        for col_idx, label in enumerate(
            ["Sitio", "Oracle", "Buk", "Diferencia", "N° Dif."], start=1
        ):
            cell = ws.cell(row=row, column=col_idx, value=label)
            cell.fill      = _HEADER_FILL
            cell.font      = _HEADER_FONT
            cell.alignment = _CENTER
        row += 1

        sitio_totals: dict[str, dict] = {}
        for rec in section_rows:
            s = rec.nom_sitio or "(sin sitio)"
            if s not in sitio_totals:
                sitio_totals[s] = {"oracle": Decimal("0"), "buk": Decimal("0"), "diffs": 0}
            sitio_totals[s]["oracle"] += rec.monto_oracle
            sitio_totals[s]["buk"]    += rec.monto_buk
            if rec.tiene_diferencia:
                sitio_totals[s]["diffs"] += 1

        for sitio, t in sorted(sitio_totals.items()):
            diff = t["oracle"] - t["buk"]
            ws.cell(row=row, column=1, value=sitio)
            for col_idx, val in [(2, float(t["oracle"])), (3, float(t["buk"])), (4, float(diff))]:
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.number_format = num_fmt
                cell.alignment     = _RIGHT
            ws.cell(row=row, column=5, value=t["diffs"])
            if abs(diff) > Decimal("0.01"):
                for col_idx in range(1, 6):
                    ws.cell(row=row, column=col_idx).fill = _RED_FILL
            row += 1

        row += 1

        # By cod_haber
        ws.cell(row=row, column=1, value="Por Código").font = _BOLD_FONT
        row += 1
        for col_idx, label in enumerate(
            ["Cód. Haber", "Descripción", "Oracle", "Buk", "Diferencia"], start=1
        ):
            cell = ws.cell(row=row, column=col_idx, value=label)
            cell.fill      = _HEADER_FILL
            cell.font      = _HEADER_FONT
            cell.alignment = _CENTER
        row += 1

        haber_totals: dict[str, dict] = {}
        for rec in section_rows:
            k = rec.cod_haber
            if k not in haber_totals:
                haber_totals[k] = {"desc": rec.descripcion_haber, "oracle": Decimal("0"), "buk": Decimal("0")}
            haber_totals[k]["oracle"] += rec.monto_oracle
            haber_totals[k]["buk"]    += rec.monto_buk

        for cod, t in sorted(haber_totals.items()):
            diff = t["oracle"] - t["buk"]
            ws.cell(row=row, column=1, value=cod)
            ws.cell(row=row, column=2, value=t["desc"])
            for col_idx, val in [(3, float(t["oracle"])), (4, float(t["buk"])), (5, float(diff))]:
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.number_format = num_fmt
                cell.alignment     = _RIGHT
            if abs(diff) > Decimal("0.01"):
                for col_idx in range(1, 6):
                    ws.cell(row=row, column=col_idx).fill = _RED_FILL
            row += 1

        row += 2  # space between sections


# ── Codes reference sheet ─────────────────────────────────────────────────────

def _build_codes_sheet(
    wb: Workbook,
    buk_to_oracle: dict[str, str],
    haber_descriptions: dict[str, str],
    rows: list[ReconciliationRow],
) -> None:
    """Sheet 'Códigos Haber' — equivalencia y estado actual Oracle cohade ↔ Buk item_code."""
    ws = wb.create_sheet("Códigos Haber")

    # ── Column definitions ────────────────────────────────────────────────────
    cols = [
        ("Oracle\ncod_haber",      16),
        ("Descripción Oracle",      30),
        ("Campo\nOracle",           13),
        ("Buk\nitem_code",          26),
        ("Campo\nBuk",              13),
        ("Tipo",                    13),
        ("N° Pers.\nOracle",        12),
        ("Total\nOracle",           14),
        ("N° Pers.\nBuk",           12),
        ("Total\nBuk",              14),
        ("Diferencia",              14),
        ("Estado",                  18),
    ]
    for col_idx, (label, width) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30

    # ── Build oracle_to_buk map ───────────────────────────────────────────────
    oracle_to_buk: dict[str, str] = {v: k for k, v in buk_to_oracle.items()}
    oracle_to_buk.update({
        "HEX051": "horas_extras_50percent",
        "HEX100": "horas_extras_100percent",
        "HEX102": "horas_extras_80percent",
        "HEX103": "horas_extras_70percent",
    })

    # ── Aggregate reconciliation data per cod_haber ───────────────────────────
    from collections import defaultdict
    stats: dict[str, dict] = defaultdict(lambda: {
        "n_oracle": 0, "total_oracle": Decimal("0"),
        "n_buk":    0, "total_buk":    Decimal("0"),
    })
    for rec in rows:
        s = stats[rec.cod_haber]
        if rec.monto_oracle:
            s["n_oracle"]    += 1
            s["total_oracle"] += rec.monto_oracle
        if rec.monto_buk:
            s["n_buk"]    += 1
            s["total_buk"] += rec.monto_buk

    # All codes = mapping table + any that appear in reconciliation but not in map
    all_codes = sorted(set(oracle_to_buk.keys()) | set(stats.keys()))

    # ── Write rows ────────────────────────────────────────────────────────────
    row = 2
    for cohade in all_codes:
        buk_code = oracle_to_buk.get(cohade, "(sin mapeo)")
        tipo     = "Hora Extra" if cohade.startswith("H") else "Bono"
        desc     = haber_descriptions.get(cohade, "")
        is_he    = cohade.startswith("H")
        num_fmt  = _HH_FMT if is_he else _AMOUNT_FMT
        s        = stats.get(cohade, {})

        n_oracle    = s.get("n_oracle", 0)
        total_oracle= s.get("total_oracle", Decimal("0"))
        n_buk       = s.get("n_buk", 0)
        total_buk   = s.get("total_buk", Decimal("0"))
        diferencia  = total_oracle - total_buk

        # Estado
        if n_oracle and n_buk:
            estado = "Con diferencia" if abs(diferencia) > Decimal("0.01") else "OK"
        elif n_oracle:
            estado = "Solo Oracle"
        elif n_buk:
            estado = "Solo Buk"
        else:
            estado = "Sin datos período"

        no_match = buk_code == "(sin mapeo)" or estado in ("Solo Oracle", "Solo Buk", "Con diferencia")
        fill = _RED_FILL if no_match and estado != "Sin datos período" else None

        def _c(col: int, value, fmt: str = "", align=_LEFT):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = align
            if fmt:
                cell.number_format = fmt
            if fill:
                cell.fill = fill
            return cell

        _c(1,  cohade)
        _c(2,  desc)
        _c(3,  "cod_haber",  align=_CENTER)
        _c(4,  buk_code)
        _c(5,  "item_code",  align=_CENTER)
        _c(6,  tipo,         align=_CENTER)
        _c(7,  n_oracle,     align=_CENTER)
        _c(8,  float(total_oracle), fmt=num_fmt, align=_RIGHT)
        _c(9,  n_buk,        align=_CENTER)
        _c(10, float(total_buk),    fmt=num_fmt, align=_RIGHT)
        _c(11, float(diferencia),   fmt=num_fmt, align=_RIGHT)
        _c(12, estado,       align=_CENTER)

        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


# ── Public entry point ────────────────────────────────────────────────────────

def generate_excel(
    rows: list[ReconciliationRow],
    periodo: str,
    output_dir: Path | None = None,
    buk_to_oracle: dict[str, str] | None = None,
    haber_descriptions: dict[str, str] | None = None,
) -> Path:
    """Generates the reconciliation Excel and returns its path."""
    if output_dir is None:
        output_dir = Path("reportes")
    output_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filepath  = output_dir / f"Cuadratura_{periodo}_{timestamp}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    _build_haberes_sheet(wb, rows)
    _build_he_sheet(wb, rows)
    _build_diff_sheet(wb, rows)
    _build_summary_sheet(wb, rows, periodo)
    if buk_to_oracle is not None:
        _build_codes_sheet(wb, buk_to_oracle, haber_descriptions or {}, rows)

    wb.save(filepath)
    logger.info("Excel guardado: %s", filepath)
    return filepath
